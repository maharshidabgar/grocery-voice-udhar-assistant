import customtkinter as ctk

from report_manager import ReportManager
from openpyxl import Workbook
from tkinter import filedialog
from datetime import datetime
from openpyxl.styles import Font, PatternFill


class ReportPage(ctk.CTkFrame):

    def __init__(self, parent):
        super().__init__(parent)

        self.report_manager = ReportManager()

        self.pack(
            fill="both",
            expand=True,
            padx=20,
            pady=20
        )

        self.create_widgets()

    # ---------------------------------------
    # Create Widgets
    # ---------------------------------------

    def create_widgets(self):

        heading = ctk.CTkLabel(
            self,
            text="📊 Reports",
            font=("Arial", 28, "bold")
        )

        heading.pack(
            pady=(10, 20)
        )

        # ---------------------------------------
        # Summary Cards
        # ---------------------------------------

        summary_frame = ctk.CTkFrame(self)

        summary_frame.pack(
            fill="x",
            padx=20,
            pady=10
        )

        summary_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)

        self.total_customer_card = ctk.CTkLabel(
            summary_frame,
            text="👥 Customers\n0",
            font=("Arial", 18, "bold"),
            width=180,
            height=80
        )

        self.total_customer_card.grid(
            row=0,
            column=0,
            padx=10,
            pady=10
        )

        self.udhar_card = ctk.CTkLabel(
            summary_frame,
            text="💰 Today's Udhar\n₹0",
            font=("Arial", 18, "bold"),
            width=180,
            height=80
        )

        self.udhar_card.grid(
            row=0,
            column=1,
            padx=10,
            pady=10
        )

        self.payment_card = ctk.CTkLabel(
            summary_frame,
            text="💵 Today's Payment\n₹0",
            font=("Arial", 18, "bold"),
            width=180,
            height=80
        )

        self.payment_card.grid(
            row=0,
            column=2,
            padx=10,
            pady=10
        )

        self.balance_card = ctk.CTkLabel(
            summary_frame,
            text="📒 Outstanding\n₹0",
            font=("Arial", 18, "bold"),
            width=180,
            height=80
        )

        self.balance_card.grid(
            row=0,
            column=3,
            padx=10,
            pady=10
        )

        # ---------------------------------------
        # Action Buttons
        # ---------------------------------------

        button_frame = ctk.CTkFrame(self)

        button_frame.pack(
            fill="x",
            padx=20,
            pady=(5, 10)
        )

        ctk.CTkButton(
            button_frame,
            text="🔄 Refresh",
            width=150,
            command=self.refresh_report
        ).pack(
            side="left",
            padx=10,
            pady=10
        )

        ctk.CTkButton(
            button_frame,
            text="📊 Export Excel",
            width=170,
            command=self.export_excel
        ).pack(
            side="left",
            padx=10,
            pady=10
        )
        # ---------------------------------------
        # Transactions List
        # ---------------------------------------

        self.report_table = ctk.CTkScrollableFrame(
            self,
            width=1000,
            height=450
        )

        self.report_table.pack(
            fill="both",
            expand=True,
            padx=20,
            pady=20
        )

        self.load_summary()

        self.load_transactions()

    # ---------------------------------------
    # Load Summary
    # ---------------------------------------

    def load_summary(self):

        total_customers = self.report_manager.get_total_customers()

        total_udhar, total_payment, total_transactions = (
            self.report_manager.get_today_summary()
        )

        outstanding = self.report_manager.get_total_outstanding()

        self.total_customer_card.configure(
            text=f"👥 Customers\n{total_customers}"
        )

        self.udhar_card.configure(
            text=f"💰 Today's Udhar\n₹{total_udhar:.2f}"
        )

        self.payment_card.configure(
            text=f"💵 Today's Payment\n₹{total_payment:.2f}"
        )

        self.balance_card.configure(
            text=f"📒 Outstanding\n₹{outstanding:.2f}"
        )

    # ---------------------------------------
    # Load Transactions
    # ---------------------------------------

    def load_transactions(self):

        for widget in self.report_table.winfo_children():

            widget.destroy()

        transactions = self.report_manager.get_today_transactions()

        if not transactions:

            ctk.CTkLabel(
                self.report_table,
                text="No Transactions Today",
                font=("Arial", 18)
            ).pack(
                pady=20
            )

            return

        # Header

        header = ctk.CTkFrame(
            self.report_table
        )

        header.pack(
            fill="x",
            pady=5
        )

        headings = [
            "Customer",
            "Type",
            "Amount",
            "Item",
            "Note",
            "Date"
        ]

        for col, title in enumerate(headings):

            ctk.CTkLabel(
                header,
                text=title,
                width=140,
                font=("Arial", 14, "bold")
            ).grid(
                row=0,
                column=col,
                padx=5
            )

        # Data

        for transaction in transactions:

            row = ctk.CTkFrame(
                self.report_table
            )

            row.pack(
                fill="x",
                pady=2
            )

            for col, value in enumerate(transaction):

                if col == 2:
                    value = f"₹{float(value):.2f}"

                ctk.CTkLabel(
                    row,
                    text=str(value),
                    width=140,
                    anchor="w"
                ).grid(
                    row=0,
                    column=col,
                    padx=5
                )

    # ---------------------------------------
    # Refresh Report
    # ---------------------------------------

    def refresh_report(self):

        self.load_summary()

        self.load_transactions()

    # ---------------------------------------
    # Export Report Excel
    # ---------------------------------------

    def export_excel(self):

        transactions = self.report_manager.get_today_transactions()

        file = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel File", "*.xlsx")],
            initialfile="Today's_Report.xlsx"
        )

        if not file:
            return

        wb = Workbook()

        ws = wb.active

        ws.title = "Today's Report"

        ws.append([
            "Customer",
            "Type",
            "Amount",
            "Item",
            "Note",
            "Date"
        ])

        header_fill = PatternFill(
            fill_type="solid",
            start_color="1F4E78"
        )

        header_font = Font(
            color="FFFFFF",
            bold=True
        )

        for cell in ws[1]:

            cell.fill = header_fill

            cell.font = header_font

        for row in transactions:

            row = list(row)

            row[2] = f"₹{float(row[2]):.2f}"

            try:

                dt = datetime.strptime(
                    row[5],
                    "%Y-%m-%d %H:%M:%S"
                )

                row[5] = dt.strftime(
                    "%d-%b-%Y %I:%M %p"
                )

            except:

                pass

            ws.append(row)

        wb.save(file)

        self.after(
            100,
            lambda: self.load_summary()
        )

        print("Report Exported :", file)